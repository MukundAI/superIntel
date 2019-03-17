import xlrd
from openpyxl import load_workbook
import csv
import os
import datetime


def read_from_excel(xlFile):
    #wb = xlrd.open_workbook(fileName)
    #sh = wb.sheet_by_name('Report')
    # Open an xlsx for reading
    wb = load_workbook(xlFile)
    # Get the current Active Sheet
    sh = wb.get_sheet_by_name("Report")
    ## iterate through the sheet to locate the date columns

    #dateIdx = 1
    #print sh.nrows
    #for rownum in sh.iter_rows():
#        dateCell = sh.cell(row = rownum, column = dateIdx).value
        #dateCell = rows[dateIdx]
        ## check if the cell is a date; continue otherwise
        #if isinstance(dateCell, float):
        #    sh.cell(row = rownum, column = dateIdx).value = xlrd.xldate_as_tuple(dateCell, wb.datemode)

        ## the "*date_tuple" will automatically unpack the tuple. Thanks mfitzp :-)

    return sh


# MAIN pulls all the Excel data from the directory
if __name__ == '__main__':
    path = 'C:\Users\mananda\Desktop\IT\EQTY\RAWtradeLog'
    CSVstorage = 'C:\ML\Stock_Market_Prediction-master\XIRR\sbiSmartdump.csv'

    xlsxFiles = [x for x in os.listdir(path) if x.endswith(".xlsx")]
    your_csv_file = open(CSVstorage, 'w')
    wr = csv.writer(your_csv_file, lineterminator='\n')  # , quoting=csv.QUOTE_ALL)

    for file in xlsxFiles:
        fullPath = os.path.join(path, file)
        data = read_from_excel(fullPath)
        wr.writerow
        #print data.cell(row=8,column=2).value
        rownum=0
        for rowval in data.iter_rows():
            rownum = rownum+1
            firstCol = data.cell(row = rownum,column = 1).value
            # print rownum.cell
            if (firstCol == 'BSE' or firstCol == 'NSE'):
                initRow=[]
                for cell in rowval:
                    initRow.append(cell.value)
                wr.writerow(initRow)

    your_csv_file.close()
