from dateutil import parser
import scipy.optimize
from datetime import time, datetime
import matplotlib as mpl
from openpyxl import load_workbook
import csv
import os
import matplotlib.pyplot as plt
from decimal import Decimal, localcontext
import numpy as np

def xnpv(rate, values, dates):
    '''Equivalent of Excel's XNPV function.

  #  >>> from datetime import date
  #  >>> dates = [date(2010, 12, 29), date(2012, 1, 25), date(2012, 3, 8)]
  #  >>> values = [-10000, 20, 10100]
  #  >>> xnpv(0.1, values, dates)
    -966.4345...
    '''
    if rate <= -1.0:
        return float('inf')
    d0 = dates[0]    # or min(dates)
    myXNPV = Decimal(0)
    for vi, di in zip(values, dates):
        with localcontext() as ctx:
            ctx.prec = 100  # 100 digits precision

        myXNPV = myXNPV + (Decimal(vi) / (Decimal(1.0) + Decimal(rate))**(Decimal((di - d0).days) / Decimal(365)))

    return myXNPV

def xirr(values, dates):
    '''Equivalent of Excel's XIRR function.


 #   >>> dates = [date(2010, 12, 29), date(2012, 1, 25), date(2012, 3, 8)]
 #   >>> values = [-10000, 20, 10100]
 #   >>> xirr(values, dates)
    0.0100612...
    '''
    try:
        return scipy.optimize.newton(lambda r: float(xnpv(r, values, dates)), 0.0)
    except RuntimeError:    # Failed to converge?
        return scipy.optimize.brentq(lambda r: xnpv(r, values, dates), -1.0, 1e10)
def cumPLOT(values,dates):
    Cval = np.cumsum(values)

    y = [np.abs(values), np.abs(Cval)]
    # Basic stacked area chart.
    plt.stackplot(dates, y, labels=['A', 'B'])
    plt.legend(loc='upper left')

    plt.show()

def mySBIplt(values,buySell,dates,stkName):

    #plt.figure(figsize=(10, 5))
    Cval = np.cumsum(values)

    y = [np.abs(values), np.abs(Cval)]
    # Basic stacked area chart.
    plt.style.use('ggplot')
    fig, ax = plt.subplots(nrows=2, sharex=True, figsize=(15, 8))

    ax[0].stackplot(dates, y, labels=['Traded price', 'Net Inflow'])
    ax[0].legend(loc='upper left')
    #pos_volm = volm
    #neg_volm = volm

    #pos_volm = [x for x in pos_volm if x > 0: return 0]
    #neg_volm[volm > 0] = 0

    ax[1].bar(dates, values, width= 1000 / (len(dates)))
    for i, stks in enumerate(stkName ):
        ax[1].annotate(stks, (dates[i], values[i]))

    for i, buyPr in enumerate(buySell ):
        ax[1].annotate(buyPr, (dates[i], values[i]+6000))

    #ax[1].bar(dates, neg_volm, width= 2000 / (len(dates)), color = 'r')
    ax[0].set_title('SBIsmart')
    ax[0].set_ylabel('Price')
    ax[1].set_ylabel('Volume')
    ax[1].get_xaxis().set_tick_params(which='major', pad=25)

    fig.autofmt_xdate()
    plt.show()
    #top = plt.subplot2grid((4, 4), (0, 0), rowspan=3, colspan=4)
    #bottom = plt.subplot2grid((4, 4), (3, 0), rowspan=1, colspan=4)
    #top.plot(dates, values, marker='o')  # CMT.index gives the dates
    #bottom.bar(dates, volm, size=4)

    # set the labels
    #xfmt = mpl.dates.DateFormatter('%Y-%D')
    #ax[1].xaxis.set_major_locator(mpl.dates.HourLocator(interval=3))
    #ax[1].xaxis.set_major_formatter(xfmt)

    #ax[1].xaxis.set_minor_locator(mpl.dates.HourLocator(interval=1))
    #ax[1].xaxis.set_minor_formatter(xfmt)





    # fig = plt.figure()
    # rect = fig.patch
    # ax1 = fig.add_subplot(1, 1, 1)
    # ax1.plot(dates, values, marker='o')
    # plt.title('EQTY-upto2017')
    # plt.xlabel('TIME')
    # plt.show()


def read_from_excel(xlFile):
        # wb = xlrd.open_workbook(fileName)
        # sh = wb.sheet_by_name('Report')
        # Open an xlsx for reading
        wb = load_workbook(xlFile)
        # Get the current Active Sheet
        sh = wb.get_sheet_by_name("Report")
        ## iterate through the sheet to locate the date columns

        # dateIdx = 1
        # print sh.nrows
        # for rownum in sh.iter_rows():
        #        dateCell = sh.cell(row = rownum, column = dateIdx).value
        # dateCell = rows[dateIdx]
        ## check if the cell is a date; continue otherwise
        # if isinstance(dateCell, float):
        #    sh.cell(row = rownum, column = dateIdx).value = xlrd.xldate_as_tuple(dateCell, wb.datemode)

        ## the "*date_tuple" will automatically unpack the tuple. Thanks mfitzp :-)

        return sh


        # MAIN pulls all the Excel data from the directory
def xlstoCSV(XLpath,CSVstorage):

    #path = 'C:\Users\mananda\Desktop\IT\EQTY\RAWtradeLog'
    #CSVstorage = 'C:\ML\Stock_Market_Prediction-master\XIRR\sbiSmartdump.csv'

    xlsxFiles = [x for x in os.listdir(XLpath) if x.endswith(".xlsx")]
    your_csv_file = open(CSVstorage, 'w')
    wr = csv.writer(your_csv_file, lineterminator='\n')  # , quoting=csv.QUOTE_ALL)

    for file in xlsxFiles:
        fullPath = os.path.join(XLpath, file)
        data = read_from_excel(fullPath)
        wr.writerow
        # print data.cell(row=8,column=2).value
        rownum = 0
        for rowval in data.iter_rows():
            rownum = rownum + 1
            firstCol = data.cell(row=rownum, column=1).value
            # print rownum.cell
            if (firstCol == 'BSE' or firstCol == 'NSE'):
                initRow = []
                for cell in rowval:
                    initRow.append(cell.value)
                wr.writerow(initRow)

    your_csv_file.close()


if __name__ =='__main__':

    #Consolidate all data in CSV
    path = 'C:\Users\mananda\Desktop\IT\EQTY\RAWtradeLog'
    CSVstorage = 'C:\ML\Stock_Market_Prediction-master\XIRR\sbiSmartdump.csv'
    #xlstoCSV(path, CSVstorage)

    #CSVstorage = 'C:\ML\Stock_Market_Prediction-master\sbiSmartdump.csv'
    values = []
    dates = []
    volm = []
    stk = []
    buySell = []
    csv_content = open(CSVstorage, 'rU')  # open the file in read universal mode


    for line in csv_content:
        cells = line.split(",")
        values.append(float(cells[6]))

        dt = parser.parse(cells[1])
        dates.append(dt)   # since we want the first, second and third column
        volm.append(int(cells[4]))
        stk.append((cells[2]))
        buySell.append((cells[5]))



    csv_content.close()
    #print values
    #sma10 = values.rolling(10).mean()
    startIdx = 40
    myxir = xirr(values[startIdx:len(values)],dates[startIdx:len(dates)])
    print myxir*100
    #print dates
    mySBIplt(values[startIdx:len(values)-1], buySell[startIdx:len(values)-1], dates[startIdx:len(values)-1], stk[startIdx:len(values)-1])

